import re
import pandas as pd

def raggruppa_per_lettera_tariffa(df):
    """
    Raggruppa i record del DataFrame per lettera/le lettere contenute nel campo 'TARIFFA'.
    Restituisce un dizionario: {lettera: DataFrame}
    """
    if 'TARIFFA' not in df.columns:
        print("Colonna 'TARIFFA' non trovata!")
        return {}
    
    # Mappatura delle categorie numeriche alle descrizioni estese
    categorie_estese = {
        '01': '01 - Edilizia',
        '02': '02 - Restauro e opere di recupero',
        '03': '03 - Infrastrutture',
        '04': '04 - Impianti elettrici',
        '05': '05 - Impianti di adduzione idrica e di scarico',
        '06': '06 - Impianti antincendio',
        '07': '07 - Impianti termici',
        '08': '08 - Fognature ed acquedotti',
        '09': '09 - Sicurezza in azienda e in cantiere',
        '10': '10 - Opere marittime',
        '11': '11 - Impianti sportivi',
        '12': '12 - Igiene ambientale',
        '13': '13 - Opere idrauliche e di bonifica e consolidamento',
        '14': '14 - Opere forestali',
        '15': '15 - Sondaggi e prove',
        '16': '16 - Opere a verde e irrigazione',
        '17': '17 - Arredo urbano e parco giochi'
    }
    
    # Estrai il numero della categoria dalla tariffa
    def estrai_categoria(tariffa):
        if pd.isna(tariffa):
            return None
        match = re.search(r'/(\d+)\.', str(tariffa))
        if match:
            numero_cat = match.group(1)
            return categorie_estese.get(numero_cat, numero_cat)
        return None
    
    df['CATEGORIA_ESTESA'] = df['TARIFFA'].apply(estrai_categoria)
    gruppi = {}
    for categoria, gruppo in df.groupby('CATEGORIA_ESTESA'):
        if categoria:
            gruppi[categoria] = gruppo
    return gruppi

def crea_struttura_gerarchica(df):
    """
    Crea una struttura gerarchica basata sui codici tariffa.
    Raggruppa per codice base (es: 01.E01.001) e tiene i dettagli come figli.
    """
    if 'TARIFFA' not in df.columns:
        return {}
    
    def estrai_codice_base(tariffa):
        if pd.isna(tariffa):
            return None
        # Estrae il pattern: PUG2025/01.E01.001
        match = re.search(r'(PUG\d+/\d+\.[A-Z]+\.\d+)', str(tariffa))
        if match:
            return match.group(1)
        return None
    
    def estrai_dettaglio(tariffa):
        if pd.isna(tariffa):
            return None
        # Estrae tutto dopo il codice base: .001, .003, etc.
        match = re.search(r'PUG\d+/\d+\.[A-Z]+\.\d+(.+)', str(tariffa))
        if match:
            return match.group(1)
        return ""
    
    df_copia = df.copy()
    df_copia['CODICE_BASE'] = df_copia['TARIFFA'].apply(estrai_codice_base)
    df_copia['DETTAGLIO'] = df_copia['TARIFFA'].apply(estrai_dettaglio)
    
    struttura = {}
    
    for codice_base, gruppo in df_copia.groupby('CODICE_BASE'):
        if codice_base:
            # Ordina per dettaglio per avere un ordine logico
            gruppo_ordinato = gruppo.sort_values('DETTAGLIO')
            
            # Il primo elemento (senza dettaglio o con dettaglio minimo) è il nodo principale
            nodo_principale = None
            dettagli = []
            
            for _, row in gruppo_ordinato.iterrows():
                if row['DETTAGLIO'] == "" or row['DETTAGLIO'] == ".001":
                    nodo_principale = row.to_dict()
                else:
                    dettagli.append(row.to_dict())
            
            # Se non c'è un nodo principale, prendi il primo
            if nodo_principale is None and len(gruppo_ordinato) > 0:
                nodo_principale = gruppo_ordinato.iloc[0].to_dict()
                dettagli = [row.to_dict() for _, row in gruppo_ordinato.iloc[1:].iterrows()]
            
            struttura[codice_base] = {
                'principale': nodo_principale,
                'dettagli': dettagli,
                'totale': len(gruppo_ordinato)
            }
    
    return struttura
import pandas as pd
import os

def esamina_excel(percorso_file):
    if not os.path.exists(percorso_file):
        print(f"File non trovato: {percorso_file}")
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(percorso_file, read_only=True, data_only=True)
        if 'Elenco Prezzi' not in wb.sheetnames:
            print("Foglio 'Elenco Prezzi' non trovato!")
            return None
        ws = wb['Elenco Prezzi']
        # Trova la riga header
        header_row = None
        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True)):
            if row and any('TARIFFA' in str(cell).upper() for cell in row if cell):
                header_row = idx
                break
        if header_row is None:
            print("Header 'TARIFFA' non trovato!")
            return None
        # Carica tutte le righe dopo l'header
        data_rows = list(ws.iter_rows(min_row=header_row+2, values_only=True))
        # Salta la prima riga se contiene 'Voce riservata'
        if data_rows and data_rows[0] and any('VOCE RISERVATA' in str(cell).upper() for cell in data_rows[0] if cell):
            data_rows = data_rows[1:]
        records = []
        i = 0
        while i < len(data_rows):
            row = data_rows[i]
            codice = str(row[1]).strip() if row and len(row) > 1 and row[1] else ''
            if codice.startswith('PUG'):
                descrizione = str(row[2]).strip() if len(row) > 2 and row[2] else ''
                # Cerca la riga successiva per unità/prezzo
                unita = ''
                prezzo = ''
                if i+1 < len(data_rows):
                    next_row = data_rows[i+1]
                    unita = str(next_row[3]).strip() if len(next_row) > 3 and next_row[3] else ''
                    prezzo = str(next_row[4]).strip() if len(next_row) > 4 and next_row[4] else ''
                records.append({
                    'TARIFFA': codice,
                    "DESCRIZIONE dell'ARTICOLO": descrizione,
                    'Unità di misura': unita,
                    'Prezzo': prezzo
                })
                i += 2
            else:
                i += 1
        df = pd.DataFrame(records)
        print(f"Colonne rilevate: {df.columns.tolist()}")
        print(df.head(10))
        return df
    except Exception as e:
        print(f"Errore durante la lettura del file: {e}")
        return None

if __name__ == "__main__":
    percorso = os.path.join("data", "regione puglia 2025.xlsx")
    df = esamina_excel(percorso)
    if df is not None:
        gruppi = raggruppa_per_lettera_tariffa(df)
        print(f"Raggruppamento per lettere LIVELLO_LETTERA trovate: {list(gruppi.keys())}")
        # Mostra un esempio per una lettera
        for lettera, gruppo in gruppi.items():
            print(f"\n--- Record per lettera {lettera} ---")
            print(gruppo.head(3))
            break
